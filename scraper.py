import requests
from bs4 import BeautifulSoup
import pandas as pd
from urllib.parse import urljoin
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
import os

base_url = "https://shiny-diski.com.ua"
n = 1
result = []

os.makedirs("images", exist_ok=True)

while n < 2:
    url = f"{base_url}/uk/tires?start={n}"
    response = requests.get(url)
    soup = BeautifulSoup(response.text, "html.parser")

    prices = soup.find_all("span", class_="standard-price js-price-sorting")
    names = soup.find_all("a", class_="product-card__title js-product-name")
    pictures = soup.find_all("img", class_="product-card__image")

    for name, price, img_tag in zip(names, prices, pictures):
        # получаем ссылку на картинку
        img_url = urljoin(base_url, img_tag["src"])

        # скачиваем и сохраняем картинку
        img_data = requests.get(img_url).content
        filename = os.path.join("images", os.path.basename(img_url.split("?")[0]))
        with open(filename, "wb") as f:
            f.write(img_data)

        result.append({
            "name": name.text.strip(),
            "price": price.text.strip(),
            "image_path": filename
        })
        print(f"добавлен элемент {name.text}")
    n+= 1

# ✅ теперь вставляем картинки в Excel через openpyxl
wb = Workbook()
ws = wb.active
ws.title = "Tires"

# заголовки
ws.append(["Название", "Цена", "Фото"])

for row, item in enumerate(result, start=2):
    ws.cell(row=row, column=1, value=item["name"])
    ws.cell(row=row, column=2, value=item["price"])

    # вставляем изображение
    img = XLImage(item["image_path"])
    img.width = 80   # уменьшаем размер
    img.height = 80
    ws.row_dimensions[row].height = 60  # высота строки
    ws.add_image(img, f"C{row}")  # вставляем в колонку C

# заголовки жирным
ws.column_dimensions["A"].width = 60
ws.column_dimensions["B"].width = 15
ws.column_dimensions["C"].width = 15

# сохраняем
wb.save("tires_with_images.xlsx")
print(f"✅ Сохранено {len(result)} элементов ")